"""
批量拆分：读取 Excel 配置，对所有"拆分完成=False"的书执行拆分，完成后标记。

前提：先运行 init_work.py 生成 Excel 配置文件，再根据需要编辑 Excel。

用法：
    python split_all.py
    python split_all.py --dry-run        # 仅列出待处理书目，不实际拆分
    python split_all.py --book "书名"    # 只处理指定书名（支持部分匹配）
"""

import argparse
import sys
from pathlib import Path

import openpyxl

from split_pdf import run_split

BASE_DIR = Path(__file__).parent
BOOKS_WORK = BASE_DIR / "books-work"
BOOKS_CONFIG_PATH = BOOKS_WORK / "books_config.xlsx"
SPLIT_CONFIG_PATH = BOOKS_WORK / "split_config.xlsx"


def read_split_config() -> dict:
    """Read global split format defaults from split_config.xlsx."""
    defaults = {"max_pages": 100, "prefix_digits": 3, "prefix_sep": "-", "folder_digits": 2}

    if not SPLIT_CONFIG_PATH.exists():
        print(f"警告：找不到 {SPLIT_CONFIG_PATH.name}，使用内置默认值")
        return defaults

    wb = openpyxl.load_workbook(SPLIT_CONFIG_PATH, read_only=True, data_only=True)
    ws = wb.active

    # Row 1 = headers, Row 2 = notes, Row 3 = values
    headers = [cell.value for cell in ws[1]]
    values_row = list(ws.iter_rows(min_row=3, max_row=3, values_only=True))
    wb.close()

    if not values_row:
        return defaults

    values = values_row[0]
    cfg = dict(zip(headers, values))

    raw_mpf = cfg.get("max_pages_per_file")
    max_pages_per_file = int(raw_mpf) if raw_mpf else None
    return {
        "max_pages":          int(cfg.get("max_pages",     defaults["max_pages"])),
        "max_pages_per_file": max_pages_per_file,
        "prefix_digits":      int(cfg.get("prefix_digits", defaults["prefix_digits"])),
        "prefix_sep":         str(cfg.get("prefix_sep",    defaults["prefix_sep"])),
        "folder_digits":      int(cfg.get("folder_digits", defaults["folder_digits"])),
    }


def read_books_config() -> tuple[openpyxl.Workbook, list[dict], dict[str, int]]:
    """
    Read books_config.xlsx.
    Returns (workbook, list_of_book_dicts, col_index_map).
    col_index_map maps column name → 1-based column index.
    """
    if not BOOKS_CONFIG_PATH.exists():
        sys.exit(
            f"错误：找不到 {BOOKS_CONFIG_PATH}。\n"
            "请先运行 python init_work.py 生成配置文件。"
        )

    wb = openpyxl.load_workbook(BOOKS_CONFIG_PATH)
    ws = wb.active

    # Row 1 = headers
    headers = {cell.value: cell.column for cell in ws[1] if cell.value}

    books = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row[0]:
            continue
        d = {ws.cell(1, col).value: val for col, val in enumerate(row, 1)}
        books.append(d)

    return wb, books, headers


def to_bool(val) -> bool:
    if isinstance(val, bool):
        return val
    if isinstance(val, (int, float)):
        return bool(val)
    if isinstance(val, str):
        return val.strip().lower() in ("true", "yes", "1", "是")
    return False


def main() -> None:
    parser = argparse.ArgumentParser(description="从 Excel 批量拆分 PDF")
    parser.add_argument("--dry-run", action="store_true", help="只列出待处理书目，不实际拆分")
    parser.add_argument("--book", type=str, default=None, help="只处理书名中含此字符串的书（部分匹配）")
    args = parser.parse_args()

    split_cfg = read_split_config()
    print(f"全局拆分格式：{split_cfg}\n")

    wb, books, col_idx = read_books_config()
    ws = wb.active

    # Find which rows correspond to which books (for updating 拆分完成)
    # Build a map: book_name → row_number (1-based)
    name_to_row: dict[str, int] = {}
    for row_no, row in enumerate(ws.iter_rows(min_row=3), start=3):
        val = row[0].value
        if val:
            name_to_row[str(val)] = row_no

    pending = [b for b in books if not to_bool(b.get("拆分完成"))]
    if args.book:
        pending = [b for b in pending if args.book in str(b.get("书名", ""))]
    done_count = len(books) - len(pending)

    print(f"共 {len(books)} 本书，已完成 {done_count} 本，待处理 {len(pending)} 本")
    if not pending:
        print("所有书本均已拆分完成。")
        return

    print("\n待处理书目：")
    for b in pending:
        level = int(b.get("split_level") or 3)
        offset = int(b.get("offset") or 0)
        print(f"  · {b['书名']}  (level={level}, offset={offset})")

    if args.dry_run:
        print("\n[dry-run] 未执行拆分。")
        return

    print()
    done_col = col_idx.get("拆分完成")
    succeeded = 0
    failed = 0

    for book in pending:
        name = str(book["书名"])
        level = int(book.get("split_level") or 3)
        offset_val = book.get("offset")
        offset = int(offset_val) if offset_val is not None else None

        print(f"{'='*60}")
        print(f"正在拆分：{name}")
        print(f"{'='*60}")
        try:
            run_split(
                name,
                level=level,
                max_pages=split_cfg["max_pages"],
                max_pages_per_file=split_cfg["max_pages_per_file"],
                prefix_digits=split_cfg["prefix_digits"],
                prefix_sep=split_cfg["prefix_sep"],
                folder_digits=split_cfg["folder_digits"],
                offset=offset,
            )
            # Mark as done in workbook
            if done_col and name in name_to_row:
                ws.cell(row=name_to_row[name], column=done_col, value=True)
            wb.save(BOOKS_CONFIG_PATH)
            succeeded += 1
            print()
        except Exception as e:
            print(f"\n  [跳过] 拆分失败：{e}\n")
            failed += 1

    print(f"\n批量拆分完成：成功 {succeeded} 本，失败 {failed} 本")


if __name__ == "__main__":
    main()
