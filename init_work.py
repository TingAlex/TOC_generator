"""
初始化工作配置：扫描 books-todo/（与 books-done/）中的 PDF，登记到 Excel 配置。

生成两个文件：
  books-work/books_config.xlsx   — 每本书一行，含进度字段 + 拆分配置
  books-work/split_config.xlsx   — 全局拆分格式默认值（如不存在则创建）

用法：
    python init_work.py
"""

from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).parent
BOOKS_WORK = BASE_DIR / "books-work"
BOOKS_TODO = BASE_DIR / "books-todo"
BOOKS_DONE = BASE_DIR / "books-done"

BOOKS_CONFIG_PATH = BOOKS_WORK / "books_config.xlsx"
SPLIT_CONFIG_PATH = BOOKS_WORK / "split_config.xlsx"

# books_config.xlsx 的列定义：(列名, 说明, 默认值)
BOOKS_COLS = [
    ("书名",           "PDF文件名（不含扩展名）",    None),
    ("offset",        "印刷页码偏移量（PDF页 = 印刷页 + offset）", 0),
    ("toc_pages",     "目录所在页（逗号分隔）",        ""),
    ("split_level",   "拆分层级（1/2/3）",            3),
    ("rendered",      "目录页已渲染为图片",            False),
    ("ocr_done",      "OCR识别完成",                  False),
    ("toc_parsed",    "目录结构已解析",                False),
    ("bookmarks_added","书签已写入PDF",               False),
    ("bookmark_count","书签数量",                      0),
    ("拆分完成",       "PDF拆分已完成",                False),
]

# split_config.xlsx 的列定义：(列名, 说明, 默认值)
SPLIT_COLS = [
    ("max_pages",          "每个文件夹的最大页数",                   100),
    ("max_pages_per_file", "单文件最大页数（超出切为_1/_2，空=不限）", None),
    ("prefix_digits",      "文件前缀序号位数（3→001-）",              3),
    ("prefix_sep",         "前缀分隔符",                             "-"),
    ("folder_digits",      "文件夹编号位数（2→01）",                  2),
]

HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(color="FFFFFF", bold=True)
NOTES_FILL  = PatternFill("solid", fgColor="D9E1F2")
NOTES_FONT  = Font(color="595959", italic=True)


def style_header(cell):
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_note(cell):
    cell.fill = NOTES_FILL
    cell.font = NOTES_FONT
    cell.alignment = Alignment(horizontal="center", wrap_text=True)


def auto_width(ws, min_width=12, max_width=40):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        width = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col_letter].width = max(min_width, min(width + 2, max_width))


def _build_skeleton_ws(ws) -> None:
    """在工作表上写好表头（第1行）+ 说明（第2行）+ 冻结，无数据行。"""
    ws.title = "书本配置"
    for col_idx, (name, note, _default) in enumerate(BOOKS_COLS, 1):
        style_header(ws.cell(row=1, column=col_idx, value=name))
        style_note(ws.cell(row=2, column=col_idx, value=note))
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 40
    ws.freeze_panes = "A3"


def ensure_books_config() -> None:
    """确保 books_config.xlsx 存在（仅表头骨架，无数据行）。
    供 registry 冷启动（新机器、Excel 尚未生成）时自动调用。"""
    BOOKS_WORK.mkdir(exist_ok=True)
    if BOOKS_CONFIG_PATH.exists():
        return
    wb = openpyxl.Workbook()
    _build_skeleton_ws(wb.active)
    wb.save(BOOKS_CONFIG_PATH)


def init_books_config(pdf_names: list[str]) -> None:
    """Create or update books_config.xlsx, adding new books without touching existing rows."""
    col_names = [c[0] for c in BOOKS_COLS]

    ensure_books_config()  # 没有则建空表骨架
    wb = openpyxl.load_workbook(BOOKS_CONFIG_PATH)
    ws = wb.active

    # Row 1 = headers, Row 2 = notes, Row 3+ = data
    existing_books: set[str] = {
        str(row[0]) for row in ws.iter_rows(min_row=3, values_only=True) if row[0]
    }

    # Append new books（新书一律从默认值起步；进度由后续流程写回）
    added = 0
    for name in pdf_names:
        if name in existing_books:
            continue

        row_data = {col: default for col, _note, default in BOOKS_COLS}
        row_data["书名"] = name
        row_data["toc_pages"] = ""
        row_data["拆分完成"] = False

        row_values = [row_data[col] for col in col_names]
        ws.append(row_values)
        added += 1
        print(f"  + 新增书本：{name}")

    auto_width(ws)
    wb.save(BOOKS_CONFIG_PATH)

    if added == 0:
        print("  books_config.xlsx 无新增（所有书本已存在）")
    else:
        print(f"  books_config.xlsx 已更新，新增 {added} 本书")


def init_split_config() -> None:
    """Create split_config.xlsx with global defaults if it doesn't exist."""
    if SPLIT_CONFIG_PATH.exists():
        print("  split_config.xlsx 已存在，跳过创建")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "拆分格式"

    col_names   = [c[0] for c in SPLIT_COLS]
    col_notes   = [c[1] for c in SPLIT_COLS]
    col_defaults = [c[2] for c in SPLIT_COLS]

    # Row 1: headers
    for col_idx, name in enumerate(col_names, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        style_header(cell)
    ws.row_dimensions[1].height = 22

    # Row 2: notes
    for col_idx, note in enumerate(col_notes, 1):
        cell = ws.cell(row=2, column=col_idx, value=note)
        style_note(cell)
    ws.row_dimensions[2].height = 40

    # Row 3: default values
    for col_idx, val in enumerate(col_defaults, 1):
        ws.cell(row=3, column=col_idx, value=val)

    ws.freeze_panes = "A3"
    auto_width(ws)
    wb.save(SPLIT_CONFIG_PATH)
    print("  split_config.xlsx 已创建（默认值）")


def main() -> None:
    BOOKS_WORK.mkdir(exist_ok=True)

    # Collect all PDF names from books-todo (and books-done for already-processed ones)
    pdf_names: list[str] = []
    seen: set[str] = set()
    for src_dir in (BOOKS_TODO, BOOKS_DONE):
        if src_dir.exists():
            for f in sorted(src_dir.glob("*.pdf")):
                name = f.stem
                # Skip split output folders disguised as PDFs (none, but just in case)
                if name not in seen:
                    seen.add(name)
                    pdf_names.append(name)

    # Also pick up books that have a books-work directory but no PDF (识别目录已做、PDF 不在 todo/done)
    for d in sorted(BOOKS_WORK.iterdir()):
        if d.is_dir() and d.name not in seen and (d / "toc_parsed.txt").exists():
            seen.add(d.name)
            pdf_names.append(d.name)

    print(f"发现 {len(pdf_names)} 本书：")
    for n in pdf_names:
        print(f"  · {n}")
    print()

    print("初始化 books_config.xlsx …")
    init_books_config(pdf_names)
    print()

    print("初始化 split_config.xlsx …")
    init_split_config()
    print()

    print(f"完成！配置文件位于：{BOOKS_WORK}")
    print(f"  · {BOOKS_CONFIG_PATH.name}")
    print(f"  · {SPLIT_CONFIG_PATH.name}")


if __name__ == "__main__":
    main()
