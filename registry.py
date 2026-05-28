"""
状态管理：以 books-work/books_config.xlsx 为主要存储。
若 Excel 不存在，自动回退到各书目录下的 state.json（兜底/迁移前兼容）。
"""

import json
from pathlib import Path

WORK_DIR    = Path("books-work")
BOOKS_CONFIG = WORK_DIR / "books_config.xlsx"

# state dict 字段 → Excel 列名
_STATE_TO_COL = {
    "offset":          "offset",
    "toc_pages":       "toc_pages",
    "rendered":        "rendered",
    "ocr_done":        "ocr_done",
    "toc_parsed":      "toc_parsed",
    "bookmarks_added": "bookmarks_added",
    "bookmark_count":  "bookmark_count",
}


# ── 内部工具 ───────────────────────────────────────────────────────────────

def _stem(book_name: str) -> str:
    return book_name.removesuffix(".pdf")


def _parse_toc_pages(val) -> list[int] | None:
    if not val:
        return None
    if isinstance(val, list):
        return [int(v) for v in val]
    try:
        result = [int(v.strip()) for v in str(val).split(",") if v.strip()]
        return result or None
    except ValueError:
        return None


def _toc_pages_to_str(val) -> str:
    if isinstance(val, list):
        return ",".join(str(v) for v in val)
    return str(val) if val else ""


# ── 公开 API ───────────────────────────────────────────────────────────────

def load() -> dict:
    """聚合所有书本状态为 {book_name.pdf: state_dict}。
    优先读 books_config.xlsx，不存在时回退到各 state.json。"""
    if BOOKS_CONFIG.exists():
        return _load_from_excel()
    return _load_from_state_json()


def save(registry: dict) -> None:
    """将书本状态持久化。
    优先写 books_config.xlsx，不存在时回退到各 state.json。"""
    if BOOKS_CONFIG.exists():
        _save_to_excel(registry)
    else:
        _save_to_state_json(registry)


# ── Excel 实现 ─────────────────────────────────────────────────────────────

def _load_from_excel() -> dict:
    import openpyxl
    wb = openpyxl.load_workbook(BOOKS_CONFIG, data_only=True, read_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    registry: dict[str, dict] = {}

    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row[0]:
            continue
        d = dict(zip(headers, row))
        book_name = str(d["书名"]) + ".pdf"
        state: dict = {}

        toc_pages = _parse_toc_pages(d.get("toc_pages"))
        if toc_pages:
            state["toc_pages"] = toc_pages

        for col_name, state_key in {v: k for k, v in _STATE_TO_COL.items()}.items():
            val = d.get(col_name)
            if val is None:
                continue
            if state_key in ("rendered", "ocr_done", "toc_parsed", "bookmarks_added"):
                if val:  # 只写 True；False/空 不写入 state（等同于"未完成"）
                    state[state_key] = bool(val)
            elif state_key == "offset":
                state["offset"] = int(val)
            elif state_key == "bookmark_count":
                state["bookmark_count"] = int(val)
            # toc_pages 已单独处理

        registry[book_name] = state

    wb.close()
    return registry


def _save_to_excel(registry: dict) -> None:
    import openpyxl
    wb = openpyxl.load_workbook(BOOKS_CONFIG)
    ws = wb.active

    headers: dict[str, int] = {
        cell.value: cell.column for cell in ws[1] if cell.value
    }
    # 书名 → 行号 映射
    existing: dict[str, int] = {}
    for row in ws.iter_rows(min_row=3):
        if row[0].value:
            existing[str(row[0].value)] = row[0].row

    for book_name, state in registry.items():
        stem = _stem(book_name)
        if stem not in existing:
            # 新书：追加一行（仅书名，其余字段后续填入）
            new_row = ws.max_row + 1
            if "书名" in headers:
                ws.cell(row=new_row, column=headers["书名"], value=stem)
            existing[stem] = new_row

        target_row = existing[stem]
        for state_key, col_name in _STATE_TO_COL.items():
            if state_key not in state or col_name not in headers:
                continue
            val = state[state_key]
            if state_key == "toc_pages":
                val = _toc_pages_to_str(val)
            ws.cell(row=target_row, column=headers[col_name], value=val)

    wb.save(BOOKS_CONFIG)


# ── state.json 兜底实现（无 Excel 时使用）────────────────────────────────

def _state_path(book_name: str) -> Path:
    return WORK_DIR / _stem(book_name) / "state.json"


def _load_from_state_json() -> dict:
    registry: dict[str, dict] = {}
    for state_file in sorted(WORK_DIR.glob("*/state.json")):
        try:
            state = json.loads(state_file.read_text(encoding="utf-8"))
            book_name = state_file.parent.name + ".pdf"
            registry[book_name] = state
        except (json.JSONDecodeError, OSError):
            pass
    return registry


def _save_to_state_json(registry: dict) -> None:
    for book_name, state in registry.items():
        path = _state_path(book_name)
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(
            json.dumps(state, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
