"""
每本书的状态读写，以 books-work/books_config.xlsx 为唯一存储。

Excel 不存在时（如新机器首次运行，books-work/ 不入库），save() 会先经
bookconfig.ensure_books_config() 建好空表骨架再写入，无需手动初始化。

state dict 字段 ←→ Excel 列名见 _STATE_TO_COL。书本主键统一为 "{书名}.pdf"。
"""

from . import paths

BOOKS_CONFIG = paths.BOOKS_CONFIG_PATH

# state dict 字段 → Excel 列名
_STATE_TO_COL = {
    "offset":          "offset",
    "toc_pages":       "toc_pages",
    "rendered":        "rendered",
    "toc_parsed":      "toc_parsed",
    "bookmarks_added": "bookmarks_added",
    "bookmark_count":  "bookmark_count",
}


def _parse_toc_pages(val) -> list[int] | None:
    if not val:
        return None
    if isinstance(val, list):
        return [int(v) for v in val]
    try:
        result = [int(v.strip()) for v in str(val).split(",") if v.strip()]
        return result or None
    except ValueError:
        return None


def _toc_pages_to_str(val) -> str:
    if isinstance(val, list):
        return ",".join(str(v) for v in val)
    return str(val) if val else ""


# ── 公开 API ───────────────────────────────────────────────────────────────

def load() -> dict:
    """聚合所有书本状态为 {书名.pdf: state_dict}。Excel 不存在则返回空。"""
    if BOOKS_CONFIG.exists():
        return _load_from_excel()
    return {}


def save(registry: dict) -> None:
    """写入 books_config.xlsx；Excel 不存在则先建空表骨架。"""
    if not BOOKS_CONFIG.exists():
        from . import bookconfig
        bookconfig.ensure_books_config()
    _save_to_excel(registry)


# ── Excel 实现 ─────────────────────────────────────────────────────────────

def _load_from_excel() -> dict:
    import openpyxl
    wb = openpyxl.load_workbook(BOOKS_CONFIG, data_only=True, read_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    registry: dict[str, dict] = {}

    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row[0]:
            continue
        d = dict(zip(headers, row))
        key = paths.book_key(str(d["书名"]))
        state: dict = {}

        toc_pages = _parse_toc_pages(d.get("toc_pages"))
        if toc_pages:
            state["toc_pages"] = toc_pages

        for col_name, state_key in {v: k for k, v in _STATE_TO_COL.items()}.items():
            val = d.get(col_name)
            if val is None:
                continue
            if state_key in ("rendered", "toc_parsed", "bookmarks_added"):
                if val:  # 只写 True；False/空 视为未完成
                    state[state_key] = bool(val)
            elif state_key == "offset":
                state["offset"] = int(val)
            elif state_key == "bookmark_count":
                state["bookmark_count"] = int(val)
            # toc_pages 已单独处理

        registry[key] = state

    wb.close()
    return registry


def _save_to_excel(registry: dict) -> None:
    import openpyxl
    wb = openpyxl.load_workbook(BOOKS_CONFIG)
    ws = wb.active

    headers: dict[str, int] = {cell.value: cell.column for cell in ws[1] if cell.value}
    existing: dict[str, int] = {
        str(row[0].value): row[0].row
        for row in ws.iter_rows(min_row=3) if row[0].value
    }

    for book_name, state in registry.items():
        s = paths.stem(book_name)
        if s not in existing:
            new_row = ws.max_row + 1
            if "书名" in headers:
                ws.cell(row=new_row, column=headers["书名"], value=s)
            existing[s] = new_row

        target_row = existing[s]
        for state_key, col_name in _STATE_TO_COL.items():
            if state_key not in state or col_name not in headers:
                continue
            val = state[state_key]
            if state_key == "toc_pages":
                val = _toc_pages_to_str(val)
            ws.cell(row=target_row, column=headers[col_name], value=val)

    wb.save(BOOKS_CONFIG)
