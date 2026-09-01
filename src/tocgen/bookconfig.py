"""
Excel 配置：books_config.xlsx（每书一行）与 split_config.xlsx（全局拆分格式）。

职责聚类：本模块掌管两张 Excel 的**模板、创建、读取**；registry.py 只管每书状态的
增量读写（save 时若 Excel 不存在会回调这里的 ensure_books_config 建骨架）。
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from . import paths

BOOKS_CONFIG_PATH = paths.BOOKS_CONFIG_PATH
SPLIT_CONFIG_PATH = paths.SPLIT_CONFIG_PATH

# books_config.xlsx 列定义：(列名, 说明, 默认值)
BOOKS_COLS = [
    ("书名",            "书的相对路径（不含扩展名），如 系列/册",         None),
    ("offset",         "印刷页码偏移量（PDF页 = 印刷页 + offset）",    0),
    ("toc_pages",      "目录所在页（逗号分隔）",                        ""),
    ("split_level",    "拆分层级（1/2/3）",                            3),
    ("rendered",       "目录页已渲染为图片",                           False),
    ("toc_parsed",     "Claude 看图识别目录完成",                       False),
    ("bookmarks_added", "书签已写入PDF",                              False),
    ("bookmark_count", "书签数量",                                    0),
    ("拆分完成",        "PDF拆分已完成",                               False),
    ("边界已判读",      "边界重叠已判读（toc-boundaries + boundary_overlap.txt）", False),
    ("边界shared数",    "判读出的 shared 边界数（0 也算已判读）",         0),
]

# split_config.xlsx 列定义：(列名, 说明, 默认值)
SPLIT_COLS = [
    ("max_pages",          "每个文件夹的最大页数",                    100),
    ("max_pages_per_file", "单文件最大页数（超出切为_1/_2，空=不限）", None),
    ("prefix_digits",      "文件前缀序号位数（3→001-）",               3),
    ("prefix_sep",         "前缀分隔符",                              "-"),
    ("folder_digits",      "文件夹编号位数（2→01）",                   2),
]

_SPLIT_DEFAULTS = {c[0]: c[2] for c in SPLIT_COLS}

HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(color="FFFFFF", bold=True)
NOTES_FILL  = PatternFill("solid", fgColor="D9E1F2")
NOTES_FONT  = Font(color="595959", italic=True)


def _style_header(cell):
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _style_note(cell):
    cell.fill = NOTES_FILL
    cell.font = NOTES_FONT
    cell.alignment = Alignment(horizontal="center", wrap_text=True)


def _auto_width(ws, min_width=12, max_width=40):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        width = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col_letter].width = max(min_width, min(width + 2, max_width))


def _build_skeleton_ws(ws) -> None:
    """写好表头（第1行）+ 说明（第2行）+ 冻结，无数据行。"""
    ws.title = "书本配置"
    for col_idx, (name, note, _default) in enumerate(BOOKS_COLS, 1):
        _style_header(ws.cell(row=1, column=col_idx, value=name))
        _style_note(ws.cell(row=2, column=col_idx, value=note))
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 40
    ws.freeze_panes = "A3"


def ensure_books_config() -> None:
    """确保 books_config.xlsx 存在（仅表头骨架）。供 registry 冷启动调用。"""
    paths.BOOKS_WORK.mkdir(exist_ok=True)
    if BOOKS_CONFIG_PATH.exists():
        return
    wb = openpyxl.Workbook()
    _build_skeleton_ws(wb.active)
    wb.save(BOOKS_CONFIG_PATH)


def init_books_config(pdf_names: list[str]) -> int:
    """创建/更新 books_config.xlsx，追加新书（不动既有行）。返回新增本数。"""
    col_names = [c[0] for c in BOOKS_COLS]

    ensure_books_config()
    wb = openpyxl.load_workbook(BOOKS_CONFIG_PATH)
    ws = wb.active

    existing_books = {
        paths.stem(str(row[0])) for row in ws.iter_rows(min_row=3, values_only=True) if row[0]
    }

    added = 0
    for name in pdf_names:
        if paths.stem(name) in existing_books:
            continue
        row_data = {col: default for col, _note, default in BOOKS_COLS}
        row_data["书名"] = name
        row_data["toc_pages"] = ""
        row_data["拆分完成"] = False
        ws.append([row_data[c] for c in col_names])
        added += 1
        print(f"  + 新增书本：{name}")

    _auto_width(ws)
    wb.save(BOOKS_CONFIG_PATH)
    print("  books_config.xlsx 无新增（所有书本已存在）" if added == 0
          else f"  books_config.xlsx 已更新，新增 {added} 本书")
    return added


def init_split_config() -> None:
    """创建 split_config.xlsx（含默认值），已存在则跳过。"""
    if SPLIT_CONFIG_PATH.exists():
        print("  split_config.xlsx 已存在，跳过创建")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "拆分格式"
    for col_idx, (name, note, default) in enumerate(SPLIT_COLS, 1):
        _style_header(ws.cell(row=1, column=col_idx, value=name))
        _style_note(ws.cell(row=2, column=col_idx, value=note))
        ws.cell(row=3, column=col_idx, value=default)
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 40
    ws.freeze_panes = "A3"
    _auto_width(ws)
    wb.save(SPLIT_CONFIG_PATH)
    print("  split_config.xlsx 已创建（默认值）")


def run_init() -> None:
    """toc-init 入口：递归扫描 books-todo/(与 books-done/) 登记新书，并建 split_config。"""
    paths.BOOKS_WORK.mkdir(exist_ok=True)

    # 递归发现：书可嵌套在系列 / 教辅文件夹下，key 是相对目录根的路径（见 paths 模块）
    pdf_names: list[str] = []
    seen: set[str] = set()
    for src_dir in (paths.BOOKS_TODO, paths.BOOKS_DONE):
        for key in paths.iter_book_pdfs(src_dir):
            if key not in seen:
                seen.add(key)
                pdf_names.append(key)
    # 已识别目录（有 toc_parsed.txt）但 PDF 不在 todo/done 的也收录
    for key in paths.iter_work_books():
        if key not in seen:
            seen.add(key)
            pdf_names.append(key)

    print(f"发现 {len(pdf_names)} 本书：")
    for n in pdf_names:
        print(f"  · {n}")
    print()

    print("初始化 books_config.xlsx …")
    init_books_config(pdf_names)
    print()
    print("初始化 split_config.xlsx …")
    init_split_config()
    print()
    print(f"完成！配置文件位于：{paths.BOOKS_WORK}")
    print(f"  · {BOOKS_CONFIG_PATH.name}")
    print(f"  · {SPLIT_CONFIG_PATH.name}")


# ── 读取（供拆分流程使用） ──────────────────────────────────────────────────

def to_bool(val) -> bool:
    if isinstance(val, bool):
        return val
    if isinstance(val, (int, float)):
        return bool(val)
    if isinstance(val, str):
        return val.strip().lower() in ("true", "yes", "1", "是")
    return False


def read_split_config() -> dict:
    """读 split_config.xlsx 的全局拆分格式默认值。"""
    defaults = {"max_pages": 100, "prefix_digits": 3, "prefix_sep": "-",
                "folder_digits": 2, "max_pages_per_file": None}
    if not SPLIT_CONFIG_PATH.exists():
        print(f"警告：找不到 {SPLIT_CONFIG_PATH.name}，使用内置默认值")
        return defaults

    wb = openpyxl.load_workbook(SPLIT_CONFIG_PATH, read_only=True, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    values_row = list(ws.iter_rows(min_row=3, max_row=3, values_only=True))
    wb.close()
    if not values_row:
        return defaults

    cfg = dict(zip(headers, values_row[0]))
    raw_mpf = cfg.get("max_pages_per_file")
    return {
        "max_pages":          int(cfg.get("max_pages",     defaults["max_pages"])),
        "max_pages_per_file": int(raw_mpf) if raw_mpf else None,
        "prefix_digits":      int(cfg.get("prefix_digits", defaults["prefix_digits"])),
        "prefix_sep":         str(cfg.get("prefix_sep",    defaults["prefix_sep"])),
        "folder_digits":      int(cfg.get("folder_digits", defaults["folder_digits"])),
    }


def read_books_config():
    """
    读 books_config.xlsx。返回 (workbook, [book_dict,…], {列名: 1-based 列号})。
    workbook 保持打开，供调用方写回（如标记 拆分完成）后 save。
    """
    if not BOOKS_CONFIG_PATH.exists():
        raise FileNotFoundError(
            f"找不到 {BOOKS_CONFIG_PATH}。请先运行 toc-init 生成配置文件。")

    wb = openpyxl.load_workbook(BOOKS_CONFIG_PATH)
    ws = wb.active
    headers = {cell.value: cell.column for cell in ws[1] if cell.value}
    books = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row[0]:
            continue
        books.append({ws.cell(1, col).value: val for col, val in enumerate(row, 1)})
    return wb, books, headers
